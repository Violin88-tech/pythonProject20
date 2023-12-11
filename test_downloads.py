import os
import zipfile
from zipfile import ZipFile
from importlib import resources
from pypdf import PdfReader
import tmp
import pypdf
from openpyxl import load_workbook


def create_zip():
    file_dir = os.listdir(tmp)
    archive_name = os.path.join(resources, 'test.zip')

    if not os.path.exists("resourses/"):
        os.mkdir("resourses/")

        with zipfile.ZipFile('test.zip', mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
            for file in file_dir:
                zf.write(os.path.join(tmp, file), file)


def test_zip():
    with ZipFile('resourses/zip.zip', "r") as zf:
        test_zip = zf.namelist()
    file_list_in_folder = os.listdir("tmp/")
    assert test_zip == file_list_in_folder


    with ZipFile.open("file.pdf") as pdf:
        reader = PdfReader(pdf)
        text = reader.pages[0].extract_text()
        assert 'pytest Documentation' in text


    with ZipFile.open("csv.csv") as csv:
        text = zf.read('csv.csv')
    assert text == b'Privet'


    with ZipFile.open("") as xlsx:
        workbook = load_workbook(xlsx)
        sheet = workbook.active
        a = sheet.cell(row=1, column=1).value
        assert a == 'Privet'

